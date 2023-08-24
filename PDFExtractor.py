from collections import Counter
from io import StringIO

import PyPDF2
import pandas as pd
from tabula import read_pdf
import os
import fitz
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer, LTTextLine, LTChar, LTImage, LTFigure
import openpyxl
from ExternalModel import ExternalModel, PromptTemplate
from OCR import OCR


class PDFExtractor:
    """A PDF extractor for extracting contents from PDF files"""

    @staticmethod
    def pipeline(*args: str):
        """Pipeline of extracting content from a PDF file"""
        with open('apikey.txt', 'r') as apikey_container:
            apikey = apikey_container.read()
        model = ExternalModel(apikey)
        for file in args:
            print("-" * 35)
            print("Extracting content from " + file)
            file_name = os.path.splitext(os.path.basename(file))[0]
            output_directory = "testcases/outputs/" + file_name
            os.makedirs(output_directory)
            # extract images from PDF file
            print("-" * 25)
            print("Extracting images")
            curr_dir = output_directory + "/images"
            os.makedirs(curr_dir)
            PDFExtractor.extract_images(file, curr_dir)
            # flag indicate if the directory is deleted
            flag = deal_with_dir(curr_dir)
            # extract text from the output images
            print("-" * 25)
            print("Extracting text from output images")
            # if there are images extracted
            if not flag:
                curr_dir = output_directory + "/text_from_images"
                os.makedirs(curr_dir)
                PDFExtractor.extract_text_from_images(
                    output_directory + "/images", curr_dir, external_model=model)
            # extract tables from PDF file
            print("-" * 25)
            print("Extracting tables")
            PDFExtractor.extract_tables(file, output_directory + "/tables.xlsx",
                                        external_model=model, set_wrap_text=True)
            # extract text from PDF file
            print("-" * 25)
            print("Extracting text")
            PDFExtractor.extract_text(file, output_directory + "/text_part.txt", external_model=model)
            print("-" * 25)
            print("Done!")
            print("-" * 35)

    @staticmethod
    def extract_text(pdf_path: str, output_path: str, external_model: ExternalModel = None):
        """Extract text from PDF file"""
        try:
            pdf_reader = PyPDF2.PdfReader(pdf_path)
            with open(output_path, 'a', encoding="utf-8") as output_file:
                # extract footnotes in the pdf file
                footnotes = PDFExtractor.extract_footnotes(pdf_path)
                print(footnotes)
                # use external model to process extracted footnotes
                if footnotes:
                    processed_footnotes = PDFExtractor.process_footnotes(footnotes.copy(), external_model=external_model)
                # iterate through pages of the pdf file
                for page_num in range(len(pdf_reader.pages)):
                    # extract text from current page
                    page = pdf_reader.pages[page_num]
                    text = page.extract_text()
                    # get footnotes of the current page
                    # if footnotes is not empty
                    if footnotes:
                        current_page = footnotes[str(page_num)]

                    # use external model to correct grammar and spelling
                    # in case of random errors when extracting text
                    if text and external_model:
                        text = external_model.get_response(text, PromptTemplate.GRAMMAR_CORRECTION)

                    # delete footnotes from the text
                    # if there are footnotes in current page
                    if footnotes and current_page:
                        for footnote in current_page:
                            if footnote in text:
                                text = text.replace(footnote, "")
                        current_page = processed_footnotes[str(page_num)]
                        if current_page in text:
                            text = text.replace(current_page, "")
                    # write into the output file
                    output_file.write(f"Page {page_num + 1}:\n{text}\n")
                    # add the footnotes
                    if footnotes and current_page:
                        output_file.write(
                            "\nFootnotes:\n" + current_page + "\n")
                    print(f"page {page_num+1} finished")
        except FileNotFoundError:
            # handle file not found error
            print("File not found")

    @staticmethod
    def extract_tables(pdf_path: str, output_path: str,
                       external_model: ExternalModel = None, set_wrap_text: bool = False):
        """Extract tables from PDF file"""
        try:
            # extract tables directly using tabula.read_pdf
            tables = read_pdf(
                pdf_path, pages='all', multiple_tables=True, guess=False, lattice=True)
            results = []
            # iterate through all tables
            for table in tables:
                # skip if got nothing
                if table.empty:
                    continue
                table = table.reset_index(drop=True)
                new_table = pd.DataFrame(columns=table.columns)
                start = 0
                indexes = []

                # iterate through each line
                for index in range(len(table)):
                    # if all cells are NaN in one line
                    # consider it as a split of rows
                    # same if reach the end of the table
                    if table.iloc[index].isna().all() or index == len(table) - 1:
                        # then merge lines into one row
                        concatenated = table.iloc[start:index + 1].apply(
                            lambda r: ' '.join(map(str, r.dropna())), axis=0)
                        new_row = pd.DataFrame([concatenated.tolist()], columns=table.columns)
                        new_table = pd.concat([new_table, new_row], ignore_index=True)
                        start = index + 1

                # assume that all cells should be filled up
                # if got cell with "" value
                # merge values in other cells in the same row with the above row
                for index, row in new_table.iterrows():
                    if (row == "").any():
                        indexes.append(index)
                        concatenated = new_table.iloc[index - 1:index + 1].apply(
                            lambda r: ' '.join(map(str, r.dropna())), axis=0)
                        new_row = pd.DataFrame([concatenated.tolist()], columns=new_table.columns)
                        new_table[index - 1:index] = new_row
                # if drop raws
                # then reset index for the table in case of random error
                if indexes:
                    new_table.drop(indexes, inplace=True)
                    new_table.reset_index(drop=True, inplace=True)

                # if externalModel is passed into the method
                if external_model:
                    # check if the headers of the table are real headers
                    cols = list(table.columns)
                    is_header = external_model.get_response("; ".join(cols), PromptTemplate.ISHEADER_CHECK) in (
                        "Yes", "Yes.")

                    # if they don't seem to be header
                    # check if it should combine with the last row of previous table
                    # or the first row of the current table
                    if not is_header:
                        combine_with_previous = False
                        # combine column names(headers) with values in the first row
                        # and then send to the external model
                        cols_and_next = [f"{col} {val}" for col, val in zip(new_table.columns, new_table.iloc[0])]
                        answer_next = [external_model.get_response(pair, PromptTemplate.CONSISTENCY_CHECK)
                                       for pair in cols_and_next]
                        combine_with_next = all([ans in ("Yes.", "Yes") for ans in answer_next])
                        # if there is previous tables and number of columns are the same
                        if results and (len(results[-1].columns) == len(new_table.columns)):
                            # combine column names(headers) with values in the last row of previous table
                            cols_and_previous = [f"{val} {col}"
                                                 for val, col in zip(results[-1].iloc[-1], new_table.columns)]
                            answer_previous = [external_model.get_response(pair, PromptTemplate.CONSISTENCY_CHECK)
                                               for pair in cols_and_previous]
                            combine_with_previous = all([ans in ("Yes.", "Yes") for ans in answer_previous])

                        # if headers need to combine with both
                        if combine_with_previous and combine_with_next:
                            triple = [f"{prev} {curr}" for prev, curr in zip(cols_and_previous, new_table.iloc[0])]
                            new_row = pd.DataFrame([triple], columns=results[-1].columns)
                            new_table.columns = results[-1].columns
                            new_table = pd.concat([results[-1].iloc[:-1], new_row, new_table[1:]], ignore_index=True)
                            results.pop()
                        # if headers need to combine with last row in previous table only
                        elif combine_with_previous:
                            new_row = pd.DataFrame([cols_and_previous], columns=results[-1].columns)
                            new_table.columns = results[-1].columns
                            new_table = pd.concat([results[-1].iloc[:-1], new_row, new_table], ignore_index=True)
                            results.pop()
                        # if headers need to combine with first row in current table only
                        elif combine_with_next:
                            new_row = pd.DataFrame([cols_and_next], columns=range(1, len(new_table.columns) + 1))
                            new_table = pd.concat([new_row, new_table[1:]], ignore_index=True)
                results.append(new_table)

            # if no table extracted
            # then skip the writing phase
            if not results:
                return
            # write tables into Excel as one table per sheet
            with pd.ExcelWriter(output_path) as writer:
                for i, table in enumerate(results):
                    table.to_excel(writer, sheet_name=f'Sheet{i}', index=False)

            # set cells attribute wrapText=True for usability
            if set_wrap_text:
                workbook = openpyxl.load_workbook(output_path)
                # loop through all cells with content in the worksheet
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
                # save the modified workbook (overwriting the original file)
                workbook.save(output_path)
        except FileNotFoundError:
            print("File not found")

    @staticmethod
    def locate_images(pdf_path: str):
        """
        Locate images in PDF file
        Returns page numbers
        """

        def extract_LTImage_from_LTFigure(page_num: int, component):
            results = []
            for item in component:
                if isinstance(item, LTImage):
                    results.append((page_num, item))
                elif isinstance(item, LTFigure):
                    results.extend(extract_LTImage_from_LTFigure(page_num, item))
            return results

        pages_contain_image = []
        for page_number, page_layout in enumerate(extract_pages(pdf_path), start=1):
            for element in page_layout:
                if isinstance(element, LTImage):
                    pages_contain_image.append(page_number)
                    break
                elif isinstance(element, LTFigure):
                    temp = extract_LTImage_from_LTFigure(page_number, element)
                    if len(temp) > 0:
                        pages_contain_image.append(page_number)
                        break
        return pages_contain_image

    @staticmethod
    def extract_images(pdf_path: str, output_folder: str):
        """Extract images from PDF file"""
        try:
            pdf_document = fitz.open(pdf_path)
            # get an array of page numbers
            # for pages containing at least one images
            page_numbers = PDFExtractor.locate_images(pdf_path)
            # iterate through those pages
            for page_number in page_numbers:
                page = pdf_document[page_number - 1]
                images = page.get_images(full=True)
                for img_index, img in enumerate(images):
                    xref = img[0]
                    base_image = pdf_document.extract_image(xref)
                    image_data = base_image["image"]
                    image_type = base_image["ext"]
                    with open(f"{output_folder}/{page_number}_{img_index + 1}.{image_type}",
                              "wb") as img_file:
                        img_file.write(image_data)
            pdf_document.close()
        except FileNotFoundError:
            print("File not found")

    @staticmethod
    def extract_text_from_images(images_container: str, output_folder: str,
                                 external_model: ExternalModel = None):
        """Extract text from images"""
        OCR.extract_in_batch(images_container, output_folder, external_model=external_model)

    @staticmethod
    def extract_footnotes(pdf_path: str, font_size_threshold_factor: float = 0.95,
                          height_threshold_factor: float = 0.2):
        """Extract footnotes from the PDF file according to its position and font size"""
        font_sizes = []
        # iterate through elements in each page
        # record font sizes
        for page_layout in extract_pages(pdf_path):
            for element in page_layout:
                if isinstance(element, LTTextContainer):
                    for text_line in element:
                        if isinstance(text_line, LTTextLine):
                            for character in text_line:
                                if isinstance(character, LTChar):
                                    font_sizes.append(character.size)
                                    break
        try:
            # calculate the most common font size as the size for body
            font_size_counter = Counter(font_sizes)
            most_common_font_size = font_size_counter.most_common(1)[0][0]
            # use a factor to set a threshold for font size of footnotes
            footnote_font_size_threshold = most_common_font_size * font_size_threshold_factor
            # create a dictionary to hold footnotes in different pages
            footnotes = {}
            # iterate again to extract footnotes
            # check if font size and position satisfy the condition
            # if so, extract that part and add into footnotes
            page_number = 0
            for page_layout in extract_pages(pdf_path):
                current_page = []
                page_height = page_layout.height
                for element in page_layout:
                    if isinstance(element, LTTextContainer):
                        for text_line in element:
                            if isinstance(text_line, LTTextLine):
                                font_size = None
                                for character in text_line:
                                    if isinstance(character, LTChar):
                                        font_size = character.size
                                        break
                                y_position = text_line.y0
                                text = text_line.get_text().strip()
                                foot_note_height_threshold = page_height * height_threshold_factor
                                if font_size and font_size <= footnote_font_size_threshold \
                                        and y_position <= foot_note_height_threshold:
                                    current_page.append(text)
                footnotes[f'{page_number}'] = current_page
                page_number += 1
            return footnotes
        except IndexError as e:
            print(f"No text through out the file; {e}")

    @staticmethod
    def process_footnotes(footnotes: dict, external_model: ExternalModel):
        """Process footnotes with external model"""
        for key, values in footnotes.items():
            if values:
                contents = " ".join(values)
                # only process those with a reasonable length
                if len(contents) >= 5:
                    response = external_model.get_response(contents, PromptTemplate.GRAMMAR_CORRECTION)
                    footnotes[key] = response
                    continue
            footnotes[key] = ""
        return footnotes

    @staticmethod
    def restructure_tables(input_files: list, output_path: str, external_model: ExternalModel):
        """
        Restructure tables from unstructured text
        ps: due to dynamic response from external model, this method cannot be generalized
        """
        content = ""
        for input_path in input_files:
            with open(input_path, "r") as file:
                content += file.read()
        response = external_model.get_response(content, PromptTemplate.TABLE_RESTRUCTURE)
        try:
            lines = list(filter(lambda line: line.startswith("\""), response.split("\n")))
            contents = "\n".join(lines)
            df = pd.read_csv(StringIO(contents))
            df.to_excel(output_path, index=False)
        except pd.errors.EmptyDataError as e:
            print(e)
        except Exception as e:
            print(type(e), e)


def deal_with_dir(dir_path: str):
    """Delete the target directory if empty"""
    if len(os.listdir(dir_path)) == 0:
        try:
            os.rmdir(dir_path)
            return True
        except OSError as e:
            print(f"Error: {e}")
    return False
